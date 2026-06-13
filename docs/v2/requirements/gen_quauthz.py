#!/usr/bin/env python3
"""Generate QuAuthz-OAuth21-PQC.xlsx master workbook."""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Colour palette ──────────────────────────────────────────────────────────
C_HDR_FILL   = "1F4E78"   # dark blue  – column headers
C_SEC_FILL   = "2E75B6"   # medium blue – section headers
C_IMPL       = "C6EFCE"   # green       – IMPLEMENTED
C_PARTIAL    = "FFEB9C"   # yellow      – PARTIAL
C_MISSING    = "FFC7CE"   # pink/red    – MISSING / PLANNED
C_NA         = "D9D9D9"   # grey        – N/A
C_TITLE      = "1F4E78"   # same dark blue for title font
C_WHITE      = "FFFFFF"
C_BLACK      = "000000"

STATUS_FILL = {
    "IMPLEMENTED": C_IMPL,
    "PARTIAL":     C_PARTIAL,
    "MISSING":     C_MISSING,
    "PLANNED":     C_MISSING,
    "N/A":         C_NA,
    "SHIPPED":     C_IMPL,
    "IN PROGRESS": C_PARTIAL,
    "NOT STARTED": C_MISSING,
    "DONE":        C_IMPL,
    "OPEN":        C_MISSING,
    "CLOSED":      C_IMPL,
}

# ── Helper factories ─────────────────────────────────────────────────────────

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=10):
    return Font(bold=True, color=C_WHITE, size=size)

def title_font(size=14):
    return Font(bold=True, color=C_TITLE, size=size)

def section_font(size=10):
    return Font(bold=True, color=C_WHITE, size=size)

def body_font(size=10):
    return Font(color=C_BLACK, size=size)

def hdr_fill(hex_col=C_HDR_FILL):
    return PatternFill("solid", fgColor=hex_col)

def status_fill(status: str):
    colour = STATUS_FILL.get(status.upper(), "FFFFFF")
    return PatternFill("solid", fgColor=colour)

def wrap_align():
    return Alignment(wrap_text=True, vertical="top")

def center_align():
    return Alignment(horizontal="center", vertical="top", wrap_text=True)

ROW_H = 45   # px  (openpyxl uses points; ~45 pt ≈ 60 px; we'll use 45 as stated)

# ── Low-level cell writers ────────────────────────────────────────────────────

def write_header_row(ws, row_num, cols, fill_colour=C_HDR_FILL):
    """Write a bold white header row."""
    fill = PatternFill("solid", fgColor=fill_colour)
    for col_idx, text in enumerate(cols, start=1):
        c = ws.cell(row=row_num, column=col_idx, value=text)
        c.font      = header_font()
        c.fill      = fill
        c.alignment = center_align()
        c.border    = thin_border()

def write_section_header(ws, row_num, text, num_cols):
    """Write a full-width section header row (merged)."""
    ws.merge_cells(
        start_row=row_num, start_column=1,
        end_row=row_num, end_column=num_cols
    )
    c = ws.cell(row=row_num, column=1, value=text)
    c.font      = section_font()
    c.fill      = hdr_fill(C_SEC_FILL)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = thin_border()
    ws.row_dimensions[row_num].height = 22

def write_data_row(ws, row_num, values, status_col_idx=None):
    """Write a data row, colouring status cell if provided."""
    ws.row_dimensions[row_num].height = ROW_H
    for col_idx, val in enumerate(values, start=1):
        c = ws.cell(row=row_num, column=col_idx, value=str(val) if val is not None else "")
        c.font      = body_font()
        c.alignment = wrap_align()
        c.border    = thin_border()
        if status_col_idx and col_idx == status_col_idx:
            c.fill = status_fill(str(val))

def set_col_widths(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

def freeze_and_filter(ws, freeze_cell, filter_range=None):
    ws.freeze_panes = freeze_cell
    if filter_range:
        ws.auto_filter.ref = filter_range

# ── TAB BUILDERS ─────────────────────────────────────────────────────────────

def build_tab00(wb):
    ws = wb.create_sheet("00 Revision Log")
    ws.sheet_view.showGridLines = True

    # Title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "QuAuthz-OAuth21-PQC.xlsx — Revision Log"
    c.font  = title_font()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = ["Rev", "Date", "Author", "Summary", "Tabs Affected", "Status"]
    write_header_row(ws, 2, cols)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:F2"

    rows = [
        ["0.1", "2026-06-13", "Paul Volosen",
         "BASELINE — Initial master workbook replacing QuAuth-OIDC-PQC.xlsx. "
         "Ports all 97 rows from Tier 0 endpoint audit; adds Tiers 1-5 requirement "
         "stubs; maps open issues #139/#150/#156-#165; standards matrix; "
         "use cases UC-01–UC-08; competitive summary; roadmap.",
         "All (00-12)", "PUBLISHED"],
    ]
    for i, row in enumerate(rows, start=3):
        write_data_row(ws, i, row, status_col_idx=6)

    set_col_widths(ws, [8, 14, 18, 60, 20, 14])


def build_tab01(wb):
    ws = wb.create_sheet("01 Executive Summary")

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "QuAuthz / sts-authority — Executive Summary"
    c.font  = title_font()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Product blurb
    ws.merge_cells("A2:G4")
    blurb = ws["A2"]
    blurb.value = (
        "Product: QuAuthz / sts-authority — an open-source, vendor-neutral OAuth 2.1 "
        "Authorization Server with native Post-Quantum Cryptography (ML-DSA-65), "
        "built on the RFC 8693 STS foundation shipped in sts-delegate-rs. "
        "It is the missing piece between 'forward the user's god-token everywhere' and "
        "a full IdP — a PQC-first AS that mints narrowly-scoped OBO tokens with act claims.\n\n"
        "What QuAuthz IS: RFC 8693 STS + OAuth 2.1 AS + PQC signing + MCP-aware delegation.\n"
        "What QuAuthz IS NOT: A general-purpose IdP, an Okta clone, an LDAP replacement, "
        "or an enterprise SSO suite."
    )
    blurb.font      = body_font(10)
    blurb.alignment = Alignment(wrap_text=True, vertical="top")
    blurb.border    = thin_border()
    ws.row_dimensions[2].height = 80

    # Tier conformance matrix
    write_section_header(ws, 5, "TIER CONFORMANCE MATRIX", 7)

    tier_cols = ["Tier", "Name", "Status", "Req Count", "Conformance %",
                 "Timeline (weeks)", "Key RFCs"]
    write_header_row(ws, 6, tier_cols)
    ws.freeze_panes = "A7"

    tiers = [
        ["Tier 0", "STS Foundation",         "SHIPPED",     97,  "87%",  "Done",  "RFC 8693, 9449, 7523, 9728"],
        ["Tier 1", "OAuth 2.1 MVP",           "PLANNED",     49,  "0%",   "8",     "RFC 9700, 6749, OIDC Core, 9207"],
        ["Tier 2", "Production Hardening",    "PLANNED",     20,  "0%",   "6",     "RFC 7009, 7662, 7591, 8707"],
        ["Tier 3", "High Security",           "PLANNED",     15,  "0%",   "6",     "RFC 9449, 9126, 9101, FAPI 2.0"],
        ["Tier 4", "Delegation / MCP",        "PLANNED",     12,  "0%",   "4",     "RFC 8693, 9068, 9728"],
        ["Tier 5", "Post-Quantum Crypto",     "PLANNED",     12,  "~20%", "8",     "FIPS 204 (ML-DSA-65), FIPS 203"],
    ]
    for i, row in enumerate(tiers, start=7):
        write_data_row(ws, i, row, status_col_idx=3)
        ws.row_dimensions[i].height = ROW_H

    # Key metrics
    write_section_header(ws, 14, "KEY METRICS", 7)
    metrics = [
        ["Total requirements (all tiers)",  "205"],
        ["Tier 0 (shipped)",                "97 (87% conformance per June 2026 audit)"],
        ["Tier 1-5 (planned)",              "108"],
        ["Open issues filed",               "12 (#139, #150, #156–#165)"],
        ["PQ algorithms supported",         "ML-DSA-65 (signing), ML-KEM-768 (planned)"],
        ["Target MVP release",              "Tier 0+1 complete (Weeks 1-8 from start)"],
    ]
    write_header_row(ws, 15, ["Metric", "Value", "", "", "", "", ""])
    for i, (metric, val) in enumerate(metrics, start=16):
        ws.row_dimensions[i].height = 30
        c1 = ws.cell(row=i, column=1, value=metric)
        c1.font = Font(bold=True, size=10)
        c1.border = thin_border()
        c1.alignment = wrap_align()
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
        c2 = ws.cell(row=i, column=2, value=val)
        c2.font = body_font()
        c2.border = thin_border()
        c2.alignment = wrap_align()

    set_col_widths(ws, [12, 28, 22, 14, 16, 18, 35])


def build_tab02(wb):
    ws = wb.create_sheet("02 Tier0 STS Foundation")

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "Tier 0 — STS Foundation (SHIPPED) — 97 Requirements"
    c.font  = title_font()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = ["Endpoint", "RFC", "Section", "Requirement",
            "Status", "Evidence (file:line)", "Fix/Issue", "Priority", "Notes"]
    write_header_row(ws, 2, cols)
    freeze_and_filter(ws, "A3", "A2:I2")

    # STATUS col = 5
    STATUS = 5

    # ── POST /token — 54 rows ─────────────────────────────────────────────
    write_section_header(ws, 3, "POST /token (54 requirements)", 9)

    token_rows = [
        # RFC 8693 token exchange
        ["POST /token", "RFC 8693", "§2.1",  "grant_type MUST be urn:ietf:params:oauth:grant-type:token-exchange", "IMPLEMENTED", "service.py:48", "", "P0", ""],
        ["POST /token", "RFC 8693", "§2.1",  "subject_token MUST be present", "IMPLEMENTED", "service.py:52", "", "P0", ""],
        ["POST /token", "RFC 8693", "§2.1",  "subject_token_type MUST be present", "IMPLEMENTED", "service.py:53", "", "P0", ""],
        ["POST /token", "RFC 8693", "§2.1",  "actor_token accepted when present", "IMPLEMENTED", "service.py:60", "", "P1", ""],
        ["POST /token", "RFC 8693", "§2.1",  "actor_token_type required when actor_token present", "IMPLEMENTED", "service.py:61", "", "P1", ""],
        ["POST /token", "RFC 8693", "§2.1",  "requested_token_type validated (not ignored)", "IMPLEMENTED", "service.py:70 (#136)", "", "P1", "Fixed in #317"],
        ["POST /token", "RFC 8693", "§2.1",  "resource parameter accepted (audience binding)", "IMPLEMENTED", "service.py:75", "", "P1", ""],
        ["POST /token", "RFC 8693", "§2.1",  "audience parameter accepted", "IMPLEMENTED", "service.py:76", "", "P1", ""],
        ["POST /token", "RFC 8693", "§2.1",  "scope parameter accepted; downscoped to intersection", "IMPLEMENTED", "core.py:112", "", "P1", ""],
        ["POST /token", "RFC 8693", "§2.2",  "issued_token_type in response (JWT AT)", "IMPLEMENTED", "service.py:190", "", "P0", ""],
        ["POST /token", "RFC 8693", "§2.2",  "access_token in response", "IMPLEMENTED", "service.py:188", "", "P0", ""],
        ["POST /token", "RFC 8693", "§2.2",  "token_type = Bearer in response", "IMPLEMENTED", "service.py:192", "", "P0", ""],
        ["POST /token", "RFC 8693", "§2.2",  "expires_in in response", "IMPLEMENTED", "service.py:194", "", "P0", ""],
        ["POST /token", "RFC 8693", "§4.1",  "act claim set to actor sub in delegated token", "IMPLEMENTED", "core.py:145", "", "P0", "Delegation not impersonation"],
        ["POST /token", "RFC 8693", "§4.1",  "sub claim preserved from subject token", "IMPLEMENTED", "core.py:140", "", "P0", ""],
        ["POST /token", "RFC 8693", "§4.1",  "act.sub = actor identity (nested)", "IMPLEMENTED", "core.py:147", "", "P0", ""],
        ["POST /token", "RFC 8693", "§4.3",  "may_act enforcement when present in subject token", "PARTIAL",      "core.py:130", "#156", "P1", "Advisory only; not enforced strictly"],
        ["POST /token", "RFC 8693", "§2.1",  "scope downscoped to subject token scopes", "IMPLEMENTED", "core.py:115", "", "P1", "RFC 9700 policy"],
        ["POST /token", "RFC 8693", "§2.3",  "invalid_request on missing required params", "IMPLEMENTED", "errors.py:28", "", "P0", ""],
        ["POST /token", "RFC 7523", "§2.2",  "client_assertion_type urn:ietf:params:oauth:client-assertion-type:jwt-bearer", "IMPLEMENTED", "client_auth.py:35", "", "P0", ""],
        ["POST /token", "RFC 7523", "§2.2",  "client_assertion JWT validated (sig, exp, aud, iss)", "IMPLEMENTED", "client_auth.py:60-95", "", "P0", ""],
        ["POST /token", "RFC 7523", "§3",    "client_id in assertion must match sub claim", "IMPLEMENTED", "client_auth.py:88", "", "P0", "RFC 7521 §4.2"],
        ["POST /token", "RFC 7523", "§3",    "assertion aud must match token endpoint URI", "IMPLEMENTED", "client_auth.py:75", "", "P0", ""],
        ["POST /token", "RFC 7523", "§3",    "assertion exp must not be expired", "IMPLEMENTED", "client_auth.py:72", "", "P0", ""],
        ["POST /token", "RFC 7523", "§3",    "assertion jti uniqueness / replay prevention", "IMPLEMENTED", "client_auth.py:95 replay.py", "", "P0", ""],
        ["POST /token", "RFC 7523", "§3",    "client_assertion sig verified against CLIENT_JWKS_FILE", "IMPLEMENTED", "client_auth.py:82", "", "P0", ""],
        ["POST /token", "RFC 7523", "§3",    "invalid_client + 401 on client_assertion failure", "IMPLEMENTED", "client_auth.py:100 errors.py", "", "P0", ""],
        ["POST /token", "RFC 7523", "§3",    "private_key_jwt: actor_token failures → invalid_request (not invalid_client)", "IMPLEMENTED", "service.py:85", "", "P0", "Invariant #6"],
        ["POST /token", "RFC 9449", "§4",    "DPoP proof validated when DPoP header present", "IMPLEMENTED", "dpop.py:45", "", "P0", "Slice 1 shipped #332"],
        ["POST /token", "RFC 9449", "§4.2",  "DPoP proof htm = POST", "IMPLEMENTED", "dpop.py:58", "", "P0", ""],
        ["POST /token", "RFC 9449", "§4.2",  "DPoP proof htu = token endpoint URI", "IMPLEMENTED", "dpop.py:62", "", "P0", ""],
        ["POST /token", "RFC 9449", "§4.2",  "DPoP proof iat freshness check (±60s)", "IMPLEMENTED", "dpop.py:70", "", "P0", ""],
        ["POST /token", "RFC 9449", "§4.2",  "DPoP proof jti uniqueness (stateless HMAC approach)", "IMPLEMENTED", "dpop.py:82", "", "P0", "Stateless §4.3"],
        ["POST /token", "RFC 9449", "§4.2",  "DPoP proof jwk header present and not private key", "IMPLEMENTED", "dpop.py:38", "", "P0", ""],
        ["POST /token", "RFC 9449", "§4.2",  "DPoP proof typ = dpop+jwt", "IMPLEMENTED", "dpop.py:32", "", "P0", ""],
        ["POST /token", "RFC 9449", "§4.3",  "cnf.jkt thumbprint bound in issued token", "PARTIAL",      "service.py:200", "#157", "P0", "Planned slice 2"],
        ["POST /token", "RFC 9449", "§5",    "token_type = DPoP in response when DPoP proof present", "PARTIAL",      "service.py:192", "#158", "P1", "Still returns Bearer"],
        ["POST /token", "RFC 9449", "§9",    "DPoP-Token-V header in response (nonce — optional)", "MISSING",     "", "#159", "P2", "Nonce not yet issued"],
        ["POST /token", "RFC 8725", "§2.1",  "alg=none rejected", "IMPLEMENTED", "verify.py:88", "", "P0", ""],
        ["POST /token", "RFC 8725", "§2.1",  "HS256/symmetric alg rejected for subject tokens", "IMPLEMENTED", "verify.py:92", "", "P0", ""],
        ["POST /token", "RFC 8725", "§3.3",  "exp claim validated on subject token", "IMPLEMENTED", "verify.py:110", "", "P0", ""],
        ["POST /token", "RFC 8725", "§3.3",  "nbf claim validated when present", "IMPLEMENTED", "verify.py:115", "", "P0", ""],
        ["POST /token", "RFC 8725", "§3.6",  "iss validated against IDP_ISSUER config", "IMPLEMENTED", "verify.py:120", "", "P0", ""],
        ["POST /token", "RFC 8725", "§3.6",  "aud validated on subject token", "IMPLEMENTED", "verify.py:125", "", "P0", ""],
        ["POST /token", "RFC 9068", "§2.2",  "iss in issued AT = STS issuer", "IMPLEMENTED", "signer.py:38", "", "P0", ""],
        ["POST /token", "RFC 9068", "§2.2",  "sub in issued AT = user sub from subject token", "IMPLEMENTED", "signer.py:42", "", "P0", ""],
        ["POST /token", "RFC 9068", "§2.2",  "aud in issued AT = requested resource/audience", "IMPLEMENTED", "signer.py:46", "", "P0", ""],
        ["POST /token", "RFC 9068", "§2.2",  "exp in issued AT = configurable TTL", "IMPLEMENTED", "signer.py:50", "", "P0", ""],
        ["POST /token", "RFC 9068", "§2.2",  "jti (unique ID) in issued AT", "IMPLEMENTED", "signer.py:55", "", "P0", ""],
        ["POST /token", "RFC 9068", "§2.2",  "scope in issued AT = granted scope", "IMPLEMENTED", "signer.py:48", "", "P0", ""],
        ["POST /token", "RFC 9068", "§4",    "subject OIDC auth context (acr/amr) carried to AT", "IMPLEMENTED", "service.py:155 (#98)", "", "P1", "Fixed #327"],
        ["POST /token", "RFC 6749", "§5.2",  "error response JSON with error field", "IMPLEMENTED", "errors.py:15", "", "P0", ""],
        ["POST /token", "RFC 6749", "§5.2",  "error_description present", "IMPLEMENTED", "errors.py:18", "", "P0", ""],
        ["POST /token", "RFC 6749", "§5.2",  "WWW-Authenticate header on 401", "IMPLEMENTED", "server.py:210", "", "P0", ""],

        # ── GET /jwks — 6 rows ───────────────────────────────────────────
    ]

    jwks_rows = [
        ["GET /jwks", "RFC 7517", "§4",    "JWKS endpoint returns public keys only", "IMPLEMENTED", "keys.py:85 server.py:240", "", "P0", ""],
        ["GET /jwks", "RFC 7517", "§4.7",  "kid present in each JWK", "IMPLEMENTED", "keys.py:90", "", "P0", ""],
        ["GET /jwks", "RFC 7517", "§4",    "kty correct for key type (EC/RSA/OKP)", "IMPLEMENTED", "keys.py:88", "", "P0", ""],
        ["GET /jwks", "RFC 7517", "§4",    "No private key material in response", "IMPLEMENTED", "keys.py:95", "", "P0", ""],
        ["GET /jwks", "RFC 8414", "§2",    "jwks_uri in AS metadata points to /jwks", "IMPLEMENTED", "server.py:260", "", "P0", ""],
        ["GET /jwks", "RFC 9449", "§5",    "jwks includes DPoP-bound key thumbprint reference (planned)", "MISSING", "", "#161", "P2", ""],
    ]

    introspect_rows = [
        ["POST /introspect", "RFC 7662", "§2.1",  "token parameter required", "IMPLEMENTED", "server.py:310", "", "P0", ""],
        ["POST /introspect", "RFC 7662", "§2.2",  "active: true/false in response", "IMPLEMENTED", "server.py:320", "", "P0", ""],
        ["POST /introspect", "RFC 7662", "§2.2",  "scope in response when active", "IMPLEMENTED", "server.py:325", "", "P0", ""],
        ["POST /introspect", "RFC 7662", "§2.2",  "sub in response when active", "IMPLEMENTED", "server.py:328", "", "P0", ""],
        ["POST /introspect", "RFC 7662", "§2.2",  "exp in response when active", "IMPLEMENTED", "server.py:330", "", "P0", ""],
        ["POST /introspect", "RFC 7662", "§2.2",  "iss in response when active", "IMPLEMENTED", "server.py:332", "", "P0", ""],
        ["POST /introspect", "RFC 7662", "§2.2",  "token_type in response", "MISSING",      "", "#162", "P1", "Filed #162"],
        ["POST /introspect", "RFC 7662", "§2.1",  "token_type_hint parameter accepted", "PARTIAL",      "server.py:312", "#164", "P2", "Accepted but ignored; filed #164"],
        ["POST /introspect", "RFC 7662", "§2.1",  "Caller must be authorized (client auth)", "PARTIAL",      "server.py:308", "#163", "P1", "Any bearer accepted; filed #163"],
        ["POST /introspect", "RFC 7662", "§2.2",  "act claim preserved in introspection response", "IMPLEMENTED", "server.py:340", "", "P1", ""],
    ]

    revoke_rows = [
        ["POST /revoke", "RFC 7009", "§2.1",  "token parameter required", "IMPLEMENTED", "server.py:380", "", "P0", ""],
        ["POST /revoke", "RFC 7009", "§2.2",  "200 OK even for unknown tokens", "IMPLEMENTED", "server.py:392", "", "P0", "RFC 7009 §2.2"],
        ["POST /revoke", "RFC 7009", "§2.1",  "token_type_hint accepted", "PARTIAL",      "server.py:382", "#164", "P2", "Accepted but not used"],
        ["POST /revoke", "RFC 7009", "§2.1",  "Caller must be authorized (only own tokens)", "MISSING",      "", "#139", "P0", "Any client can revoke any token; filed #139"],
        ["POST /revoke", "RFC 7009", "§2.2",  "invalid_client on auth failure", "IMPLEMENTED", "server.py:388", "", "P0", ""],
        ["POST /revoke", "RFC 7009", "§2",    "Revocation is persistent (replay store integration)", "PARTIAL",      "replay.py", "#26", "P1", "Per-process only; multi-worker gap"],
        ["POST /revoke", "RFC 7009", "§2.1",  "unsupported_token_type error when applicable", "N/A",          "", "", "P3", "Only issues one token type"],
        ["POST /revoke", "RFC 7009", "§2.2",  "Token family revocation (refresh tokens)", "N/A",          "", "", "P3", "No refresh tokens yet"],
    ]

    wellknown_rows = [
        ["GET /.well-known/oauth-authorization-server", "RFC 8414", "§2",    "issuer exact match (no trailing slash)", "IMPLEMENTED", "server.py:260", "", "P0", ""],
        ["GET /.well-known/oauth-authorization-server", "RFC 8414", "§2",    "token_endpoint present", "IMPLEMENTED", "server.py:262", "", "P0", ""],
        ["GET /.well-known/oauth-authorization-server", "RFC 8414", "§2",    "jwks_uri present", "IMPLEMENTED", "server.py:264", "", "P0", ""],
        ["GET /.well-known/oauth-authorization-server", "RFC 8414", "§2",    "grant_types_supported includes token-exchange", "IMPLEMENTED", "server.py:266", "", "P0", ""],
        ["GET /.well-known/oauth-authorization-server", "RFC 8414", "§2",    "token_endpoint_auth_methods_supported", "IMPLEMENTED", "server.py:270", "", "P0", ""],
        ["GET /.well-known/oauth-authorization-server", "RFC 8414", "§2",    "scopes_supported present", "IMPLEMENTED", "server.py:268", "", "P1", ""],
        ["GET /.well-known/oauth-authorization-server", "RFC 8414", "§2",    "response_types_supported present", "N/A",          "", "", "P2", "No /authorize yet"],
        ["GET /.well-known/oauth-authorization-server", "RFC 9449", "§5",    "dpop_signing_alg_values_supported present", "PARTIAL",      "server.py:275", "#165", "P1", "Field present but incomplete"],
        ["GET /.well-known/oauth-authorization-server", "RFC 8414", "§2",    "introspection_endpoint present", "IMPLEMENTED", "server.py:272", "", "P1", ""],
        ["GET /.well-known/oauth-authorization-server", "RFC 8414", "§2",    "revocation_endpoint present", "IMPLEMENTED", "server.py:274", "", "P1", ""],
        ["GET /.well-known/oauth-authorization-server", "RFC 8693", "§2",    "token_exchange grant advertised in metadata", "IMPLEMENTED", "server.py:266", "", "P0", ""],
        ["GET /.well-known/oauth-authorization-server", "RFC 8414", "§2",    "service_documentation optional but included", "IMPLEMENTED", "server.py:280", "", "P3", ""],
    ]

    pr_rows = [
        ["GET /.well-known/oauth-protected-resource", "RFC 9728", "§2",    "resource field = resource URI", "IMPLEMENTED", "server.py:295", "", "P0", "Already shipped"],
        ["GET /.well-known/oauth-protected-resource", "RFC 9728", "§2",    "authorization_servers list present", "IMPLEMENTED", "server.py:297", "", "P0", ""],
        ["GET /.well-known/oauth-protected-resource", "RFC 9728", "§2",    "bearer_methods_supported present", "IMPLEMENTED", "server.py:299", "", "P0", ""],
        ["GET /.well-known/oauth-protected-resource", "RFC 9728", "§2",    "scopes_supported present", "IMPLEMENTED", "server.py:301", "", "P1", ""],
        ["GET /.well-known/oauth-protected-resource", "RFC 9728", "§2",    "jwks_uri present for RS key publication", "IMPLEMENTED", "server.py:303", "", "P0", ""],
        ["GET /.well-known/oauth-protected-resource", "RFC 9728", "§2",    "resource_signing_alg_values_supported", "PARTIAL",      "server.py:305", "#160", "P1", "alg not user-configurable; filed #160"],
    ]

    mcp_rows = [
        ["ANY /mcp", "RFC 9728", "§",     "Bearer token required on all MCP requests", "IMPLEMENTED", "server.py:350", "", "P0", ""],
        ["ANY /mcp", "RFC 8693", "§4.1",  "act claim present on delegated MCP tokens", "IMPLEMENTED", "core.py:145", "", "P0", ""],
        ["ANY /mcp", "RFC 9449", "§7",    "DPoP token validation at MCP resource server (guidance)", "PARTIAL",      "server.py:355", "#150", "P1", "RS guidance documented; enforcement partial"],
    ]

    all_data = token_rows + jwks_rows + introspect_rows + revoke_rows + wellknown_rows + pr_rows + mcp_rows

    # Section boundaries (row num in ws, 1-indexed; data starts at row 4)
    section_markers = {
        len(token_rows) + 4:            "GET /jwks (6 requirements)",
        len(token_rows) + len(jwks_rows) + 5: "POST /introspect (10 requirements)",
        len(token_rows) + len(jwks_rows) + len(introspect_rows) + 6: "POST /revoke (8 requirements)",
        len(token_rows) + len(jwks_rows) + len(introspect_rows) + len(revoke_rows) + 7: "GET /.well-known/oauth-authorization-server (12 requirements)",
        len(token_rows) + len(jwks_rows) + len(introspect_rows) + len(revoke_rows) + len(wellknown_rows) + 8: "GET /.well-known/oauth-protected-resource (6 requirements)",
        len(token_rows) + len(jwks_rows) + len(introspect_rows) + len(revoke_rows) + len(wellknown_rows) + len(pr_rows) + 9: "ANY /mcp (3 requirements)",
    }

    row = 4
    for group_rows, section_name in [
        (token_rows,     "POST /token (54 requirements)"),
        (jwks_rows,      "GET /jwks (6 requirements)"),
        (introspect_rows,"POST /introspect (10 requirements)"),
        (revoke_rows,    "POST /revoke (8 requirements)"),
        (wellknown_rows, "GET /.well-known/oauth-authorization-server (12 requirements)"),
        (pr_rows,        "GET /.well-known/oauth-protected-resource (6 requirements)"),
        (mcp_rows,       "ANY /mcp (3 requirements)"),
    ]:
        if row > 4:
            write_section_header(ws, row, section_name, 9)
            row += 1
        for r in group_rows:
            write_data_row(ws, row, r, status_col_idx=5)
            row += 1

    set_col_widths(ws, [28, 12, 10, 55, 15, 28, 10, 10, 35])


def build_tab03(wb):
    ws = wb.create_sheet("03 Tier1 OAuth21 MVP")

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "Tier 1 — OAuth 2.1 MVP (BUILD FIRST) — 49 Requirements"
    c.font  = title_font()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = ["Req ID", "Endpoint", "RFC", "Section", "Requirement",
            "Tier", "Status", "Acceptance Evidence (curl shape)", "Priority", "Notes"]
    write_header_row(ws, 2, cols)
    freeze_and_filter(ws, "A3", "A2:J2")

    groups = {
        "GET/POST /authorize — 15 requirements": [
            ["T1-AZ-01", "GET/POST /authorize", "RFC 9700",   "§4.3.1", "HTTPS only; reject HTTP redirect URIs",                        "Tier 1", "PLANNED", "curl http:// → 400 invalid_request", "P0", ""],
            ["T1-AZ-02", "GET/POST /authorize", "RFC 9700",   "§4.3.2", "PKCE required (code_challenge + code_challenge_method=S256)",   "Tier 1", "PLANNED", "curl without code_challenge → 400", "P0", ""],
            ["T1-AZ-03", "GET/POST /authorize", "RFC 6749",   "§4.1.1", "response_type=code (only; no implicit)",                        "Tier 1", "PLANNED", "response_type=token → 400", "P0", ""],
            ["T1-AZ-04", "GET/POST /authorize", "RFC 6749",   "§4.1.1", "redirect_uri must match registered URI exactly",                "Tier 1", "PLANNED", "unregistered redirect_uri → 400", "P0", ""],
            ["T1-AZ-05", "GET/POST /authorize", "RFC 6749",   "§4.1.1", "state parameter MUST be echoed back",                          "Tier 1", "PLANNED", "state in → state out in redirect", "P0", ""],
            ["T1-AZ-06", "OIDC Core",           "§1",         "§3.1.2", "nonce accepted for OIDC flows; echoed in ID token",             "Tier 1", "PLANNED", "nonce in → nonce in id_token", "P1", ""],
            ["T1-AZ-07", "GET/POST /authorize", "RFC 9207",   "§2",     "iss parameter included in authorization response",              "Tier 1", "PLANNED", "redirect includes &iss=<issuer>", "P0", ""],
            ["T1-AZ-08", "GET/POST /authorize", "RFC 6749",   "§4.1.2", "auth code issued (single-use, short-lived ~60s)",               "Tier 1", "PLANNED", "code present in redirect URI", "P0", ""],
            ["T1-AZ-09", "GET/POST /authorize", "RFC 9700",   "§4.3.2", "auth code is one-time-use; replay returns invalid_grant",       "Tier 1", "PLANNED", "replay code → 400 invalid_grant", "P0", ""],
            ["T1-AZ-10", "GET/POST /authorize", "RFC 6749",   "§4.1.2", "scope from request reflected in code (session binding)",        "Tier 1", "PLANNED", "scope reflected in token response", "P1", ""],
            ["T1-AZ-11", "GET/POST /authorize", "RFC 6749",   "§4.1.2.1","error=access_denied redirect when user denies",               "Tier 1", "PLANNED", "deny → redirect with error=access_denied", "P1", ""],
            ["T1-AZ-12", "GET/POST /authorize", "RFC 6749",   "§4.1.2.1","error=invalid_scope if scope not supported",                  "Tier 1", "PLANNED", "bogus scope → redirect with error=invalid_scope", "P1", ""],
            ["T1-AZ-13", "GET/POST /authorize", "RFC 6749",   "§4.1.2.1","error=invalid_request for malformed requests",                "Tier 1", "PLANNED", "missing response_type → error redirect", "P0", ""],
            ["T1-AZ-14", "GET/POST /authorize", "RFC 9700",   "§4.1.2", "No credentials in query string or fragment",                    "Tier 1", "PLANNED", "verify no token in redirect query", "P0", ""],
            ["T1-AZ-15", "GET/POST /authorize", "RFC 6749",   "§4.1.1", "client_id validated against client registry",                  "Tier 1", "PLANNED", "unknown client_id → 400 invalid_client", "P0", ""],
        ],
        "POST /token (auth_code grant) — 10 requirements": [
            ["T1-TK-01", "POST /token", "RFC 6749",  "§4.1.3", "grant_type=authorization_code accepted",                                "Tier 1", "PLANNED", "grant_type=authorization_code → 200", "P0", ""],
            ["T1-TK-02", "POST /token", "RFC 9700",  "§4.3.2", "code_verifier validated against code_challenge (PKCE)",                 "Tier 1", "PLANNED", "wrong verifier → 400 invalid_grant", "P0", ""],
            ["T1-TK-03", "POST /token", "RFC 6749",  "§4.1.3", "redirect_uri must match value used in /authorize",                     "Tier 1", "PLANNED", "mismatched redirect_uri → 400", "P0", ""],
            ["T1-TK-04", "POST /token", "RFC 7523",  "§2",     "Client auth required (private_key_jwt or client_secret)",              "Tier 1", "PLANNED", "no client auth → 401 invalid_client", "P0", ""],
            ["T1-TK-05", "POST /token", "RFC 6749",  "§4.1.3", "auth code one-time-use; replay → invalid_grant",                       "Tier 1", "PLANNED", "replay code → 400 invalid_grant", "P0", ""],
            ["T1-TK-06", "POST /token", "RFC 9068",  "§2.2",   "Access token issued as AT+JWT (JWT profile)",                          "Tier 1", "PLANNED", "at+jwt typ in token header", "P0", ""],
            ["T1-TK-07", "OIDC Core",   "§3.1.3.3", "§",      "ID token issued when openid scope requested",                           "Tier 1", "PLANNED", "openid scope → id_token in response", "P0", ""],
            ["T1-TK-08", "OIDC Core",   "§3.1.3.3", "§",      "ID token contains sub, iss, aud, exp, iat, nonce",                      "Tier 1", "PLANNED", "decode id_token and verify claims", "P0", ""],
            ["T1-TK-09", "RFC 6749",    "§5.1",      "§",      "refresh_token optional; issued when offline_access scope",             "Tier 1", "PLANNED", "offline_access → refresh_token present", "P2", ""],
            ["T1-TK-10", "RFC 6749",    "§5.1",      "§",      "scope in token response reflects granted scope",                       "Tier 1", "PLANNED", "scope in response matches or is subset", "P1", ""],
        ],
        "GET /userinfo — 6 requirements": [
            ["T1-UI-01", "GET /userinfo", "OIDC Core", "§5.3",  "Bearer token required",                                                "Tier 1", "PLANNED", "no token → 401", "P0", ""],
            ["T1-UI-02", "GET /userinfo", "OIDC Core", "§5.3",  "sub claim in response must match access token sub",                    "Tier 1", "PLANNED", "sub matches token sub", "P0", ""],
            ["T1-UI-03", "GET /userinfo", "OIDC Core", "§5.4",  "Claims gated by scope (email only if email scope, etc.)",             "Tier 1", "PLANNED", "no email scope → no email in response", "P1", ""],
            ["T1-UI-04", "GET /userinfo", "OIDC Core", "§5.3",  "No over-disclosure of claims beyond granted scope",                   "Tier 1", "PLANNED", "only scoped claims returned", "P0", ""],
            ["T1-UI-05", "GET /userinfo", "OIDC Core", "§5.3",  "Content-Type: application/json",                                      "Tier 1", "PLANNED", "check response Content-Type", "P1", ""],
            ["T1-UI-06", "GET /userinfo", "OIDC Core", "§5.3",  "act claim included when delegation token used",                       "Tier 1", "PLANNED", "delegated token → act in userinfo", "P1", ""],
        ],
        "GET /.well-known/openid-configuration — 8 requirements": [
            ["T1-OD-01", "GET /.well-known/openid-configuration", "OIDC Discovery", "§3", "issuer exact match (no trailing slash)",     "Tier 1", "PLANNED", "issuer == BASE_URL exactly", "P0", ""],
            ["T1-OD-02", "GET /.well-known/openid-configuration", "OIDC Discovery", "§3", "authorization_endpoint present",             "Tier 1", "PLANNED", "field present and reachable", "P0", ""],
            ["T1-OD-03", "GET /.well-known/openid-configuration", "OIDC Discovery", "§3", "token_endpoint present",                     "Tier 1", "PLANNED", "field present", "P0", ""],
            ["T1-OD-04", "GET /.well-known/openid-configuration", "OIDC Discovery", "§3", "jwks_uri present",                           "Tier 1", "PLANNED", "field present and 200", "P0", ""],
            ["T1-OD-05", "GET /.well-known/openid-configuration", "OIDC Discovery", "§3", "userinfo_endpoint present",                  "Tier 1", "PLANNED", "field present", "P1", ""],
            ["T1-OD-06", "GET /.well-known/openid-configuration", "OIDC Discovery", "§3", "response_types_supported includes code",     "Tier 1", "PLANNED", "code in list", "P0", ""],
            ["T1-OD-07", "GET /.well-known/openid-configuration", "OIDC Discovery", "§3", "subject_types_supported present",            "Tier 1", "PLANNED", "field present", "P1", ""],
            ["T1-OD-08", "GET /.well-known/openid-configuration", "OIDC Discovery", "§3", "id_token_signing_alg_values_supported",      "Tier 1", "PLANNED", "field present, ES256 or ML-DSA-65", "P0", ""],
        ],
        "Client Registry — 5 requirements": [
            ["T1-CR-01", "Client Registry", "RFC 6749",  "§2",   "client_id unique and stable",                                        "Tier 1", "PLANNED", "client_id collision rejected", "P0", ""],
            ["T1-CR-02", "Client Registry", "RFC 6749",  "§2",   "redirect_uri registration enforced (exact match)",                   "Tier 1", "PLANNED", "unregistered URI rejected", "P0", ""],
            ["T1-CR-03", "Client Registry", "RFC 6749",  "§2",   "grant_types registered per client",                                  "Tier 1", "PLANNED", "wrong grant_type → invalid_client", "P1", ""],
            ["T1-CR-04", "Client Registry", "RFC 7523",  "§2",   "token_endpoint_auth_method registered per client",                   "Tier 1", "PLANNED", "wrong auth method → invalid_client", "P0", ""],
            ["T1-CR-05", "Client Registry", "RFC 7517",  "§4",   "JWKS or jwks_uri registered for private_key_jwt clients",           "Tier 1", "PLANNED", "no JWKS → 400 on client auth", "P0", ""],
        ],
        "Session Subsystem — 5 requirements": [
            ["T1-SS-01", "Session", "RFC 6749",  "§4.1",  "Auth code bound to session (PKCE + redirect_uri)",                          "Tier 1", "PLANNED", "different redirect_uri → invalid_grant", "P0", ""],
            ["T1-SS-02", "Session", "RFC 9700",  "§4.3",  "Auth code one-time-use (replay store)",                                     "Tier 1", "PLANNED", "second use of same code → invalid_grant", "P0", ""],
            ["T1-SS-03", "Session", "RFC 6749",  "§4.1",  "Auth code expiry (≤60 seconds)",                                            "Tier 1", "PLANNED", "expired code → invalid_grant", "P0", ""],
            ["T1-SS-04", "Session", "OIDC Core", "§3.1",  "Session state bound to nonce in ID token",                                  "Tier 1", "PLANNED", "nonce mismatch → reject", "P1", ""],
            ["T1-SS-05", "Session", "RFC 9700",  "§4.3",  "No session fixation: new session on login",                                 "Tier 1", "PLANNED", "old session cookie invalidated", "P0", ""],
        ],
    }

    row = 3
    for section_name, reqs in groups.items():
        write_section_header(ws, row, section_name, 10)
        row += 1
        for r in reqs:
            write_data_row(ws, row, r, status_col_idx=7)
            row += 1

    set_col_widths(ws, [12, 32, 14, 10, 55, 8, 12, 45, 10, 25])


def build_tab04(wb):
    ws = wb.create_sheet("04 Tier2 Production")

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "Tier 2 — Production Hardening — ~20 Requirements"
    c.font  = title_font()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = ["Req ID", "Endpoint", "RFC", "Section", "Requirement",
            "Tier", "Status", "Acceptance Evidence (curl shape)", "Priority", "Notes"]
    write_header_row(ws, 2, cols)
    freeze_and_filter(ws, "A3", "A2:J2")

    groups = {
        "RFC 9700 Security BCP — 5 requirements": [
            ["T2-SEC-01", "ALL",        "RFC 9700", "§4.1.3", "HTTPS-only; HTTP → 301 or 400",                                         "Tier 2", "PLANNED", "curl http:// → redirect or 400", "P0", ""],
            ["T2-SEC-02", "POST /token","RFC 9700", "§4.3.1", "Rate limiting on token endpoint (≤100 req/min per client)",              "Tier 2", "PLANNED", "exceed limit → 429 Too Many Requests", "P0", ""],
            ["T2-SEC-03", "GET /auth",  "RFC 9700", "§4.2.1", "Implicit flow not supported (response_type=token rejected)",             "Tier 2", "PLANNED", "response_type=token → 400", "P0", ""],
            ["T2-SEC-04", "ALL",        "RFC 9700", "§4.1.3", "Exact redirect_uri match; no wildcard or partial match",                 "Tier 2", "PLANNED", "partial match → 400 invalid_request", "P0", ""],
            ["T2-SEC-05", "ALL",        "RFC 9700", "§4.1.3", "No credentials/tokens in query string",                                  "Tier 2", "PLANNED", "inspect redirect URI for token in query", "P0", ""],
        ],
        "RFC 7009 Revocation Hardening — 3 requirements": [
            ["T2-RV-01", "POST /revoke","RFC 7009", "§2.1",   "Cross-client check: client can only revoke own tokens",                  "Tier 2", "PLANNED", "client A cannot revoke client B token → 200 (silent deny)", "P0", "Fix #139"],
            ["T2-RV-02", "POST /revoke","RFC 7009", "§2",     "Token family revocation when refresh tokens exist",                      "Tier 2", "PLANNED", "revoke refresh → access tokens in family also invalid", "P2", ""],
            ["T2-RV-03", "POST /revoke","RFC 7009", "§2",     "Revocation persists across restarts (durable store)",                    "Tier 2", "PLANNED", "revoke → restart → token still inactive", "P0", "Fix #26 equivalent"],
        ],
        "RFC 7662 Introspection Hardening — 3 requirements": [
            ["T2-IX-01", "POST /introspect","RFC 7662","§2.2", "token_type field included in introspection response",                   "Tier 2", "PLANNED", "active response has token_type field", "P1", "Fix #162"],
            ["T2-IX-02", "POST /introspect","RFC 7662","§2.1", "token_type_hint used to optimize lookup",                              "Tier 2", "PLANNED", "hint speeds lookup (benchmark test)", "P2", "Fix #164"],
            ["T2-IX-03", "POST /introspect","RFC 7662","§2.1", "Caller authorization enforced (registered RS only)",                   "Tier 2", "PLANNED", "unregistered caller → 401", "P1", "Fix #163"],
        ],
        "RFC 7591 Dynamic Client Registration — 3 requirements": [
            ["T2-DCR-01","POST /register","RFC 7591","§3.1",  "client_name, redirect_uris, grant_types in registration",               "Tier 2", "PLANNED", "POST /register → 201 with client_id", "P2", "Future"],
            ["T2-DCR-02","POST /register","RFC 7591","§3.2",  "client_secret or JWKS issued in response",                              "Tier 2", "PLANNED", "response has credentials", "P2", ""],
            ["T2-DCR-03","POST /register","RFC 7591","§3.1",  "redirect_uri scheme validation (https only)",                           "Tier 2", "PLANNED", "http redirect_uri → 400", "P1", ""],
        ],
        "RFC 8707 Resource Indicators — 2 requirements": [
            ["T2-RI-01", "POST /token","RFC 8707","§2",       "resource parameter scopes token audience",                              "Tier 2", "PLANNED", "aud in AT = requested resource", "P1", ""],
            ["T2-RI-02", "POST /token","RFC 8707","§2",       "Multiple resource parameters accepted (one AT per resource)",           "Tier 2", "PLANNED", "2 resource params → 2 calls or error", "P2", ""],
        ],
        "Infrastructure — 4 requirements": [
            ["T2-INF-01","ALL",         "Infra",    "§",      "Multi-replica replay store (Redis/Postgres backed)",                    "Tier 2", "PLANNED", "2 replicas share jti store", "P0", "Fix #26"],
            ["T2-INF-02","GET /jwks",   "RFC 7517", "§4",     "Key rotation with JWKS overlap window",                                "Tier 2", "PLANNED", "old kid valid during overlap", "P0", "Fix #25"],
            ["T2-INF-03","POST /token", "RFC 9449", "§",      "DPoP nonce issuance and rotation",                                     "Tier 2", "PLANNED", "server nonce in DPoP-Nonce header", "P1", "Fix #159"],
            ["T2-INF-04","ALL",         "Infra",    "§",      "Structured audit log (actor, subject, action, outcome)",               "Tier 2", "PLANNED", "JSON audit trail per token exchange", "P0", ""],
        ],
    }

    row = 3
    for section_name, reqs in groups.items():
        write_section_header(ws, row, section_name, 10)
        row += 1
        for r in reqs:
            write_data_row(ws, row, r, status_col_idx=7)
            row += 1

    set_col_widths(ws, [12, 22, 12, 10, 55, 8, 12, 45, 10, 20])


def build_tab05(wb):
    ws = wb.create_sheet("05 Tier3 High Security")

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "Tier 3 — High Security — ~15 Requirements"
    c.font  = title_font()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = ["Req ID", "Endpoint", "RFC", "Section", "Requirement",
            "Tier", "Status", "Acceptance Evidence (curl shape)", "Priority", "Notes"]
    write_header_row(ws, 2, cols)
    freeze_and_filter(ws, "A3", "A2:J2")

    groups = {
        "RFC 9449 DPoP — Required by Default — 5 requirements": [
            ["T3-DP-01","POST /token","RFC 9449","§4",    "DPoP required by default (no fallback to Bearer)",                        "Tier 3", "PLANNED", "no DPoP header → 400 use_dpop_nonce", "P0", "Tier 3 default-on policy"],
            ["T3-DP-02","POST /token","RFC 9449","§4.3",  "cnf.jkt binding enforced in issued token",                                "Tier 3", "PLANNED", "resource server validates cnf.jkt", "P0", "Fix #157"],
            ["T3-DP-03","ANY /resource","RFC 9449","§7",  "DPoP token validated at resource server (RS guidance included)",           "Tier 3", "PLANNED", "RS example validates DPoP proof", "P1", "Fix #150"],
            ["T3-DP-04","POST /token","RFC 9449","§9",    "Server-issued nonce required; client must retry on use_dpop_nonce",       "Tier 3", "PLANNED", "retry with DPoP-Nonce header → success", "P0", "Fix #159"],
            ["T3-DP-05","ALL",        "RFC 9449","§5",    "dpop_signing_alg_values_supported complete in metadata",                  "Tier 3", "PLANNED", "metadata lists ES256, RS256, ML-DSA-65", "P1", "Fix #165"],
        ],
        "RFC 9126 PAR — Pushed Authorization Requests — 3 requirements": [
            ["T3-PAR-01","POST /par","RFC 9126","§2",     "PAR endpoint accepts authorization parameters",                           "Tier 3", "PLANNED", "POST /par → request_uri response", "P0", ""],
            ["T3-PAR-02","GET /auth", "RFC 9126","§4",    "request_uri used in /authorize instead of inline params",                 "Tier 3", "PLANNED", "?request_uri=urn:... → auth flow", "P0", ""],
            ["T3-PAR-03","POST /par", "RFC 9126","§2",    "request_uri one-time-use; expires in 60s",                               "Tier 3", "PLANNED", "replay request_uri → 400", "P0", ""],
        ],
        "RFC 9101 JAR — JWT-Secured Auth Requests — 3 requirements": [
            ["T3-JAR-01","GET /auth","RFC 9101","§4",     "request parameter accepted (JWT-secured request object)",                 "Tier 3", "PLANNED", "?request=<JWT> → normal auth flow", "P0", ""],
            ["T3-JAR-02","GET /auth","RFC 9101","§4",     "Request JWT signature validated against client JWKS",                    "Tier 3", "PLANNED", "tampered JWT → 400 invalid_request_object", "P0", ""],
            ["T3-JAR-03","GET /auth","RFC 9101","§6",     "request_uri + request object (PAR + JAR combo) supported",               "Tier 3", "PLANNED", "PAR + JAR combo flow works", "P1", ""],
        ],
        "FAPI 2.0 Profile Alignment — 2 requirements": [
            ["T3-FAPI-01","ALL",     "FAPI 2.0","§5",     "FAPI 2.0 Baseline: PAR + PKCE + DPoP all required",                     "Tier 3", "PLANNED", "all three must pass for FAPI2 profile", "P0", ""],
            ["T3-FAPI-02","ALL",     "FAPI 2.0","§6",     "FAPI 2.0 Advanced: JAR + PAR + JARM response mode",                     "Tier 3", "PLANNED", "JAR-secured request with PAR works", "P2", ""],
        ],
        "mTLS Client Auth RFC 8705 — 2 requirements": [
            ["T3-MTLS-01","POST /token","RFC 8705","§2",   "tls_client_auth: client cert CN validated as client_id",                "Tier 3", "PLANNED", "cert CN = client_id in POST /token", "P2", "Future"],
            ["T3-MTLS-02","POST /token","RFC 8705","§2",   "self_signed_tls_client_auth method supported",                          "Tier 3", "PLANNED", "self-signed cert with pinned key works", "P2", "Future"],
        ],
    }

    row = 3
    for section_name, reqs in groups.items():
        write_section_header(ws, row, section_name, 10)
        row += 1
        for r in reqs:
            write_data_row(ws, row, r, status_col_idx=7)
            row += 1

    set_col_widths(ws, [12, 22, 12, 10, 55, 8, 12, 45, 10, 25])


def build_tab06(wb):
    ws = wb.create_sheet("06 Tier4 Delegation MCP")

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "Tier 4 — Delegation / MCP — ~12 Requirements"
    c.font  = title_font()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = ["Req ID", "Endpoint", "RFC", "Section", "Requirement",
            "Tier", "Status", "Acceptance Evidence (curl shape)", "Priority", "Notes"]
    write_header_row(ws, 2, cols)
    freeze_and_filter(ws, "A3", "A2:J2")

    groups = {
        "RFC 8693 Token Exchange wired to AS — 4 requirements": [
            ["T4-TX-01","POST /token","RFC 8693","§2.1",   "grant_type=token-exchange wired into AS /token endpoint",               "Tier 4", "PLANNED", "STS exchange via AS /token", "P0", "Port from sts-delegate-rs"],
            ["T4-TX-02","POST /token","RFC 8693","§4.1",   "act claim preserved in AS → STS chain",                                 "Tier 4", "PLANNED", "act.sub present in issued AT", "P0", "Invariant: delegation not impersonation"],
            ["T4-TX-03","POST /token","RFC 8693","§4.3",   "may_act strictly enforced (not advisory)",                              "Tier 4", "PLANNED", "actor not in may_act → 400 invalid_request", "P1", "Fix #156"],
            ["T4-TX-04","POST /token","RFC 8693","§4.1",   "Impersonation mode explicitly rejected",                                "Tier 4", "PLANNED", "no request_type=impersonation support", "P0", "Architecture invariant"],
        ],
        "RFC 9728 Protected Resource Metadata — 2 requirements": [
            ["T4-PR-01","GET /.well-known/oauth-protected-resource","RFC 9728","§2", "Already shipped; wire to AS discovery",       "Tier 4", "IMPLEMENTED", "200 with resource metadata", "P0", "Tier 0 shipped; wiring needed"],
            ["T4-PR-02","GET /.well-known/oauth-protected-resource","RFC 9728","§2", "AS cross-references PR metadata",             "Tier 4", "PLANNED", "AS metadata includes protected_resources", "P1", ""],
        ],
        "MCP-Aware Audience Policy — 3 requirements": [
            ["T4-MCP-01","ANY /mcp",  "RFC 8693","§4",    "MCP tool calls require delegated AT with act claim",                     "Tier 4", "PLANNED", "AT without act → 401 on MCP endpoints", "P0", ""],
            ["T4-MCP-02","ANY /mcp",  "RFC 9728","§",     "MCP resource server validates audience matches tool URI",                "Tier 4", "PLANNED", "wrong aud → 403", "P0", ""],
            ["T4-MCP-03","ANY /mcp",  "RFC 8707","§2",    "MCP audience scoped per tool (one token per tool server)",               "Tier 4", "PLANNED", "broad AT rejected at tool endpoint", "P1", ""],
        ],
        "RFC 9068 JWT AT Profile for Delegated Tokens — 3 requirements": [
            ["T4-AT-01","POST /token","RFC 9068","§2.2",   "act claim in JWT AT follows RFC 9068 profile",                          "Tier 4", "PLANNED", "jwt decode shows act.sub", "P0", ""],
            ["T4-AT-02","POST /token","RFC 9068","§2.2",   "client_id in JWT AT for actor identity",                                "Tier 4", "PLANNED", "client_id claim in AT", "P1", ""],
            ["T4-AT-03","POST /token","RFC 9068","§4",     "auth_time and acr carried from subject OIDC session",                   "Tier 4", "IMPLEMENTED", "auth_time in issued AT (#327)", "P1", "Fixed #327"],
        ],
    }

    row = 3
    for section_name, reqs in groups.items():
        write_section_header(ws, row, section_name, 10)
        row += 1
        for r in reqs:
            write_data_row(ws, row, r, status_col_idx=7)
            row += 1

    set_col_widths(ws, [12, 32, 12, 10, 55, 8, 14, 45, 10, 25])


def build_tab07(wb):
    ws = wb.create_sheet("07 Tier5 PQC")

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "Tier 5 — Post-Quantum Cryptography — ~12 Requirements"
    c.font  = title_font()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = ["Req ID", "Endpoint", "Standard", "Section", "Requirement",
            "Tier", "Status", "Acceptance Evidence", "Priority", "Notes"]
    write_header_row(ws, 2, cols)
    freeze_and_filter(ws, "A3", "A2:J2")

    groups = {
        "ML-DSA-65 Signing — 4 requirements": [
            ["T5-PQ-01","POST /token","FIPS 204","§5",     "ML-DSA-65 for access token signing (already in STS)",                   "Tier 5", "IMPLEMENTED","signer produces ML-DSA-65 AT", "P0", "Shipped in sts-delegate-rs"],
            ["T5-PQ-02","POST /token","FIPS 204","§5",     "ML-DSA-65 for ID token signing",                                        "Tier 5", "PLANNED",    "id_token alg = ML-DSA-65", "P0", "Extend from AT"],
            ["T5-PQ-03","POST /token","RFC 7523","§2",     "ML-DSA-65 for client assertions (already in STS)",                      "Tier 5", "IMPLEMENTED","client_assertion ML-DSA-65 accepted", "P0", "Shipped in sts-delegate-rs"],
            ["T5-PQ-04","GET /jwks", "FIPS 204","§",      "ML-DSA-65 public key in JWKS (crydi5 or custom)",                        "Tier 5", "IMPLEMENTED","JWKS contains ML-DSA key", "P0", "Shipped in sts-delegate-rs"],
        ],
        "Subject Token Signing Alg Config — 1 requirement": [
            ["T5-PQ-05","POST /token","RFC 9449","§4",     "Subject token signing alg configurable (not hard-coded)",               "Tier 5", "PARTIAL",    "SIGNING_ALG env var accepted", "P1", "Fix #160; alg not user-configurable"],
        ],
        "ML-KEM-768 Key Exchange — 2 requirements": [
            ["T5-KEM-01","TLS",       "FIPS 203","§",      "ML-KEM-768 hybrid TLS (X25519MLKEM768) on all endpoints",              "Tier 5", "PLANNED",    "TLS handshake shows hybrid KEM", "P0", "No decrypt oracle per ADR"],
            ["T5-KEM-02","TLS",       "FIPS 203","§",      "No oracle risk: server decrypts only own ciphertexts",                  "Tier 5", "PLANNED",    "no plaintext decryption API exposed", "P0", "Oracle risk ADR"],
        ],
        "Hybrid Signing — 2 requirements": [
            ["T5-HY-01","POST /token","IETF Hybrid","§",   "Hybrid AT: classical (ES256) + ML-DSA-65 dual signature",               "Tier 5", "PLANNED",    "AT verifiable by classical OR PQ verifier", "P1", "Migration path for non-PQ RSs"],
            ["T5-HY-02","GET /jwks", "IETF Hybrid","§",   "JWKS exposes both EC and ML-DSA keys",                                  "Tier 5", "PLANNED",    "JWKS has 2 keys: EC + ML-DSA", "P1", ""],
        ],
        "PQC JOSE Algorithm Registration — 2 requirements": [
            ["T5-JOSE-01","GET /jwks","IANA JOSE","§",     "ML-DSA-65 alg value registered / noted in metadata",                   "Tier 5", "PARTIAL",    "alg value documented", "P1", "Fix #165 includes this"],
            ["T5-JOSE-02","ALL",      "IANA JOSE","§",     "Fallback to classical alg when PQ not supported by client",             "Tier 5", "PLANNED",    "negotiation via dpop_signing_alg_values", "P2", ""],
        ],
        "Oracle Risk Controls — 1 requirement": [
            ["T5-OR-01","ALL",        "ADR",       "§",    "No decryption oracle: KEM private key never used for decryption API",   "Tier 5", "PLANNED",    "no /decrypt endpoint; key is sign-only", "P0", "ADR controls"],
        ],
    }

    row = 3
    for section_name, reqs in groups.items():
        write_section_header(ws, row, section_name, 10)
        row += 1
        for r in reqs:
            write_data_row(ws, row, r, status_col_idx=7)
            row += 1

    set_col_widths(ws, [12, 18, 14, 10, 55, 8, 14, 45, 10, 30])


def build_tab08(wb):
    ws = wb.create_sheet("08 Open Issues Map")

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "Open Issues Map — Filed GitHub Issues to Tier/Requirement"
    c.font  = title_font()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = ["Issue #", "Title", "Tier", "Req ID", "Priority", "Status", "Notes"]
    write_header_row(ws, 2, cols)
    freeze_and_filter(ws, "A3", "A2:G2")

    issues = [
        ["#26",  "Shared replay store: per-process jti fails under multi-worker",        "Tier 2", "T2-INF-01", "P0", "OPEN",  "INF gap; multi-worker launch now fails loud (interim guard)"],
        ["#139", "POST /revoke: no cross-client authorization check",                    "Tier 2", "T2-RV-01",  "P0", "OPEN",  "Any client can revoke any token; silent security hole"],
        ["#150", "DPoP RS validation guidance missing / partial",                         "Tier 3", "T3-DP-03",  "P1", "OPEN",  "Resource server DPoP validation example/docs needed"],
        ["#156", "may_act advisory-only: not enforced as MUST",                           "Tier 4", "T4-TX-03",  "P1", "OPEN",  "RFC 8693 §4.3; fix requires strict enforcement"],
        ["#157", "cnf.jkt not bound in issued token (DPoP slice 2)",                     "Tier 1", "T1-TK-06",  "P0", "OPEN",  "DPoP sender-constraint incomplete; planned slice 2"],
        ["#158", "token_type still Bearer when DPoP proof present",                      "Tier 1", "T1-TK-06",  "P1", "OPEN",  "Should return token_type=DPoP per RFC 9449 §5"],
        ["#159", "DPoP nonce not issued by server (DPoP-Token-V header missing)",        "Tier 2", "T2-INF-03", "P2", "OPEN",  "RFC 9449 §9; optional but important for security"],
        ["#160", "resource_signing_alg / subject token alg not user-configurable",       "Tier 5", "T5-PQ-05",  "P1", "OPEN",  "SIGNING_ALG env var needed; fix #160"],
        ["#161", "JWKS missing DPoP-bound key thumbprint reference",                     "Tier 0", "—",         "P2", "OPEN",  "RFC 9449 §5 metadata gap"],
        ["#162", "POST /introspect: token_type field missing from response",             "Tier 2", "T2-IX-01",  "P1", "OPEN",  "RFC 7662 §2.2 MUST include token_type"],
        ["#163", "POST /introspect: caller authorization not enforced",                  "Tier 2", "T2-IX-03",  "P1", "OPEN",  "Any bearer token can introspect any token"],
        ["#164", "token_type_hint accepted but ignored in /introspect and /revoke",      "Tier 2", "T2-IX-02",  "P2", "OPEN",  "RFC 7662 §2.1 / RFC 7009 §2.1; hint should optimize lookup"],
        ["#165", "dpop_signing_alg_values_supported incomplete in AS metadata",          "Tier 3", "T3-DP-05",  "P1", "OPEN",  "Field present but doesn't list all supported algs"],
    ]

    for i, row_data in enumerate(issues, start=3):
        write_data_row(ws, i, row_data, status_col_idx=6)

    set_col_widths(ws, [10, 55, 10, 14, 10, 12, 45])


def build_tab09(wb):
    ws = wb.create_sheet("09 Standards Matrix")

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Standards / RFC Matrix — All Referenced Standards"
    c.font  = title_font()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = ["RFC / Standard", "Title", "Class", "Tier", "Key Requirements", "Status"]
    write_header_row(ws, 2, cols)
    freeze_and_filter(ws, "A3", "A2:F2")

    standards = [
        ["RFC 6749",     "OAuth 2.0 Authorization Framework",              "OAuth",      "Tier 1", "auth code grant, token response, error codes", "PARTIAL"],
        ["RFC 9700",     "OAuth 2.0 Security Best Current Practice",        "Security",   "Tier 2", "HTTPS-only, PKCE required, no implicit, redirect exact match", "PARTIAL"],
        ["RFC 8693",     "OAuth 2.0 Token Exchange",                        "Token Exch", "Tier 0", "grant type, subject/actor tokens, act claim", "IMPLEMENTED"],
        ["RFC 7523",     "JWT Profile for OAuth 2.0 Client Authentication", "Client Auth","Tier 0", "private_key_jwt, client assertion validation", "IMPLEMENTED"],
        ["RFC 9449",     "OAuth 2.0 Demonstrating Proof of Possession (DPoP)", "DPoP",   "Tier 0", "proof validation, cnf.jkt binding, nonce", "PARTIAL"],
        ["RFC 8414",     "OAuth 2.0 Authorization Server Metadata",         "Discovery",  "Tier 0", "/.well-known/oauth-authorization-server, required fields", "IMPLEMENTED"],
        ["RFC 9728",     "OAuth 2.0 Protected Resource Metadata",           "Discovery",  "Tier 0", "/.well-known/oauth-protected-resource", "IMPLEMENTED"],
        ["RFC 7009",     "OAuth 2.0 Token Revocation",                      "Revocation", "Tier 2", "revoke endpoint, cross-client check, silent 200", "PARTIAL"],
        ["RFC 7662",     "OAuth 2.0 Token Introspection",                   "Introspect", "Tier 2", "active flag, claims in response, caller auth", "PARTIAL"],
        ["RFC 7591",     "OAuth 2.0 Dynamic Client Registration",           "Client Reg", "Tier 2", "POST /register, client_id, redirect_uri, JWKS", "PLANNED"],
        ["RFC 8707",     "Resource Indicators for OAuth 2.0",               "Audience",   "Tier 2", "resource parameter, audience-scoped tokens", "PARTIAL"],
        ["RFC 9068",     "JWT Profile for OAuth 2.0 Access Tokens",         "AT Profile", "Tier 0", "AT+JWT claims: iss, sub, aud, exp, jti, scope", "IMPLEMENTED"],
        ["RFC 7517",     "JSON Web Key (JWK)",                              "JOSE",       "Tier 0", "JWKS format, kid, kty, no private key material", "IMPLEMENTED"],
        ["RFC 8725",     "JWT Best Current Practices",                      "Security",   "Tier 0", "alg=none rejected, symmetric alg rejected, exp/nbf", "IMPLEMENTED"],
        ["RFC 7521",     "Assertion Framework for OAuth 2.0",               "Client Auth","Tier 0", "assertion sub = client_id, aud = token endpoint", "IMPLEMENTED"],
        ["RFC 9126",     "OAuth 2.0 Pushed Authorization Requests (PAR)",   "High Sec",   "Tier 3", "POST /par, request_uri, one-time-use", "PLANNED"],
        ["RFC 9101",     "JWT-Secured Authorization Request (JAR)",         "High Sec",   "Tier 3", "request parameter, JWT validation, PAR+JAR", "PLANNED"],
        ["RFC 9207",     "OAuth 2.0 Authorization Server Issuer Identification", "Security","Tier 1","iss in auth response", "PLANNED"],
        ["RFC 8705",     "Mutual-TLS Client Authentication",                "Client Auth","Tier 3", "tls_client_auth, self_signed_tls_client_auth", "PLANNED"],
        ["OIDC Core",    "OpenID Connect Core 1.0",                         "OIDC",       "Tier 1", "ID token, userinfo, nonce, sub claim", "PLANNED"],
        ["OIDC Discovery","OpenID Connect Discovery 1.0",                   "OIDC",       "Tier 1", "/.well-known/openid-configuration, required fields", "PLANNED"],
        ["FAPI 2.0",     "Financial-grade API Security Profile 2.0",        "Profile",    "Tier 3", "PAR + PKCE + DPoP required, JAR advanced", "PLANNED"],
        ["FIPS 204",     "ML-DSA (Module Lattice Digital Signature)",        "PQC",        "Tier 5", "ML-DSA-65 signing for AT, ID tokens, assertions", "PARTIAL"],
        ["FIPS 203",     "ML-KEM (Module Lattice Key Encapsulation Mechanism)","PQC",     "Tier 5", "ML-KEM-768 hybrid TLS, no oracle", "PLANNED"],
    ]

    for i, row_data in enumerate(standards, start=3):
        write_data_row(ws, i, row_data, status_col_idx=6)

    set_col_widths(ws, [16, 50, 14, 10, 55, 15])


def build_tab10(wb):
    ws = wb.create_sheet("10 Use Cases")

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "QuAuthz Use Cases — UC-01 through UC-08"
    c.font  = title_font()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = ["UC ID", "Title", "Tier", "Actor", "Scenario",
            "Key Requirements", "Status", "Notes"]
    write_header_row(ws, 2, cols)
    freeze_and_filter(ws, "A3", "A2:H2")

    use_cases = [
        ["UC-01", "AI Agent OBO Token Exchange",
         "Tier 0+4", "AI agent (actor) + user (subject)",
         "User logs in via Okta/IdP. AI agent presents user's broad token to STS, "
         "receives narrowly-scoped AT with act={sub:agent} bound to one MCP tool server.",
         "RFC 8693 §4.1, act claim, scope downscoping, RFC 9068 AT profile",
         "IMPLEMENTED", "Core sts-delegate-rs use case; shipped"],
        ["UC-02", "PKCE Auth Code Flow (Browser App)",
         "Tier 1", "End user (browser)",
         "SPA redirects user to /authorize with PKCE. User authenticates. "
         "App exchanges code at /token for AT+JWT + refresh token.",
         "RFC 9700 PKCE, RFC 6749 §4.1, RFC 9068 AT+JWT, OIDC ID token",
         "PLANNED", "New in Tier 1; requires /authorize + session subsystem"],
        ["UC-03", "MCP Gateway Delegation",
         "Tier 0+4", "MCP gateway (actor) + user (subject)",
         "Gateway presents user's broad AT to STS. STS mints MCP-audience-scoped "
         "AT with act claim. Gateway forwards to MCP tool server.",
         "RFC 8693, RFC 9728 resource metadata, RFC 8707 resource indicators",
         "PARTIAL", "Token exchange shipped; audience policy wiring in Tier 4"],
        ["UC-04", "DPoP Sender-Constrained Token",
         "Tier 0+3", "Client with DPoP key",
         "Client sends DPoP proof at /token. STS validates proof and binds cnf.jkt "
         "in issued AT. Resource server validates DPoP proof on each request.",
         "RFC 9449 §4, cnf.jkt, token_type=DPoP, RS validation guidance",
         "PARTIAL", "Proof validation shipped (#332); cnf.jkt binding pending (#157)"],
        ["UC-05", "FAPI 2.0 High-Security API Access",
         "Tier 3", "Regulated API client",
         "Client pushes auth request via PAR. Uses JAR JWT-secured request. "
         "Receives code, exchanges with DPoP proof for AT. FAPI 2.0 profile enforced.",
         "RFC 9126 PAR, RFC 9101 JAR, RFC 9449 DPoP, FAPI 2.0",
         "PLANNED", "Tier 3 new build"],
        ["UC-06", "Post-Quantum Signed Tokens",
         "Tier 5", "PQC-capable client",
         "QuAuthz signs AT + ID tokens with ML-DSA-65. Client verifies with "
         "PQ public key from /jwks. Hybrid mode supports classical verifiers in parallel.",
         "FIPS 204 ML-DSA-65, /jwks with PQ key, hybrid signing",
         "PARTIAL", "AT signing shipped in STS; ID token PQ signing planned"],
        ["UC-07", "Dynamic Client Registration",
         "Tier 2", "New API client",
         "Client POSTs to /register with redirect_uris, grant_types, JWKS. "
         "Receives client_id. Can immediately request authorization.",
         "RFC 7591 DCR, client registry, HTTPS validation",
         "PLANNED", "Tier 2 future; requires client registry store"],
        ["UC-08", "Token Introspection by Resource Server",
         "Tier 2", "Resource server",
         "RS calls POST /introspect with opaque token. Receives active=true/false, "
         "scope, sub, act, exp. Acts on response without decoding AT locally.",
         "RFC 7662, token_type in response, caller auth enforcement",
         "PARTIAL", "Endpoint shipped; token_type (#162) + caller auth (#163) gaps"],
    ]

    for i, row_data in enumerate(use_cases, start=3):
        write_data_row(ws, i, row_data, status_col_idx=7)

    set_col_widths(ws, [8, 30, 10, 25, 60, 45, 14, 30])


def build_tab11(wb):
    ws = wb.create_sheet("11 Competitive Summary")

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "QuAuthz — Competitive Summary"
    c.font  = title_font()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = ["Dimension", "Okta", "Auth0", "Keycloak", "AWS Cognito", "QuAuthz", "QuAuthz Advantage"]
    write_header_row(ws, 2, cols)
    freeze_and_filter(ws, "A3", "A2:G2")

    rows = [
        ["Cost",                    "$2–$8/user/mo",    "$0.023/MAU",       "Free (self-host)",     "$0.0055/MAU",  "Free / OSS (self-host)", "Zero per-user cost; no vendor lock-in"],
        ["Local-first / air-gap",   "No (cloud-only)",  "No (cloud-only)",  "Yes (Docker)",         "No (AWS)",     "Yes (single binary)",    "Runs anywhere: laptop, edge, container, air-gapped"],
        ["Passwordless default",    "Optional (add-on)","Optional",         "Optional (plugin)",    "Optional",     "Token-exchange native",  "No passwords in the delegation path at all"],
        ["OAuth 2.1 aligned",       "Partial",          "Partial",          "Partial",              "Partial",      "Yes (design goal)",      "First-class OAuth 2.1 with BCP compliance by design"],
        ["DPoP support",            "Partial (preview)","Partial",          "Partial",              "No",           "Yes (Tier 0+3)",         "DPoP built-in; required-by-default at Tier 3"],
        ["PQC-ready",               "No",               "No",               "No (roadmap vague)",   "No",           "Yes (ML-DSA-65)",        "Only AS with ML-DSA-65 + ML-KEM-768 roadmap"],
        ["RFC 8693 act claim",      "No (impersonation only)","No",          "No",                   "No",           "Yes (core invariant)",   "True delegation (sub+act), not impersonation"],
        ["MCP-aware",               "No",               "No",               "No",                   "No",           "Yes (Tier 4)",           "Native MCP audience policy + act-claim enforcement"],
        ["Open source",             "No",               "No",               "Yes (Apache 2.0)",     "No",           "Yes (MIT/Apache)",       "Fully auditable; no proprietary blobs"],
        ["STS token exchange",      "Enterprise only",  "Enterprise only",  "Plugin required",      "No",           "Core feature",           "RFC 8693 is the foundation, not a bolt-on"],
        ["Audit trail",             "Yes (SIEM add-on)","Yes (add-on)",     "Yes (events)",         "Yes (CloudTrail)","Planned (Tier 2)",    "Built-in structured audit log per exchange"],
        ["Hybrid PQ+classical sign","No",               "No",               "No",                   "No",           "Planned (Tier 5)",       "Hybrid signing enables gradual PQ migration"],
    ]

    for i, row_data in enumerate(rows, start=3):
        write_data_row(ws, i, row_data, status_col_idx=None)
        # highlight QuAuthz column (col 6)
        c6 = ws.cell(row=i, column=6)
        c6.fill = PatternFill("solid", fgColor=C_IMPL)
        c6.font = Font(bold=True, size=10)

    set_col_widths(ws, [25, 18, 18, 18, 16, 20, 45])


def build_tab12(wb):
    ws = wb.create_sheet("12 Roadmap")

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "QuAuthz Roadmap — Phase / Milestone Plan"
    c.font  = title_font()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    cols = ["Phase", "Tier", "Weeks", "Deliverables",
            "Conformance Target", "Key RFCs", "Risk", "Mitigation"]
    write_header_row(ws, 2, cols)
    freeze_and_filter(ws, "A3", "A2:H2")

    phases = [
        ["Phase 0\n(DONE)",
         "Tier 0", "Done",
         "RFC 8693 STS shipped:\n• POST /token (token exchange)\n"
         "• GET /jwks, POST /introspect, POST /revoke\n"
         "• GET /.well-known/oauth-authorization-server\n"
         "• GET /.well-known/oauth-protected-resource\n"
         "• DPoP proof validation (RFC 9449 stateless, #332)\n"
         "• ML-DSA-65 signing\n"
         "• private_key_jwt client auth\n"
         "• act claim delegation (invariant #1)",
         "87% (70/80 reqs per June 2026 audit)",
         "RFC 8693, 9449, 7523, 9068, 8414, 9728, 8725",
         "12 open issues (#139, #150, #156-#165)",
         "File and prioritize; fix P0 gaps before Tier 1 freeze"],
        ["Phase 1\n(MVP)",
         "Tier 1", "8",
         "OAuth 2.1 Authorization Code flow:\n"
         "• GET/POST /authorize (PKCE, state, iss RFC 9207)\n"
         "• POST /token (auth_code grant, PKCE verify)\n"
         "• GET /userinfo (scope-gated claims)\n"
         "• GET /.well-known/openid-configuration\n"
         "• Client Registry (client_id, redirect_uri, JWKS)\n"
         "• Session subsystem (auth code, one-time-use, expiry)",
         "Tier 0+1: ~95% (fix P0 open issues)",
         "RFC 9700, 6749, OIDC Core, 9207, 8414",
         "Session store complexity; PKCE state machine",
         "Start with in-memory store; add Redis adapter in Phase 2"],
        ["Phase 2\n(Production)",
         "Tier 2", "6",
         "Production hardening:\n"
         "• RFC 9700 BCP enforcement (HTTPS, rate limit, exact redirect)\n"
         "• RFC 7009 revocation cross-client check (fix #139)\n"
         "• RFC 7662 introspection: token_type + caller auth (fix #162/#163)\n"
         "• RFC 7591 Dynamic Client Registration\n"
         "• Multi-replica replay store (fix #26)\n"
         "• Key rotation + JWKS overlap window\n"
         "• Structured audit log",
         "Tier 0+1+2: ~98%",
         "RFC 9700, 7009, 7662, 7591, 8707",
         "Durable store migration; replay store distributed coordination",
         "Use Redis with Lua scripts for atomic jti check; document migration path"],
        ["Phase 3\n(High Sec)",
         "Tier 3", "6",
         "High-security features:\n"
         "• RFC 9449 DPoP required by default (fix #157/#158/#159)\n"
         "• RFC 9126 PAR endpoint\n"
         "• RFC 9101 JAR support\n"
         "• FAPI 2.0 Baseline profile\n"
         "• mTLS client auth (RFC 8705, initial)",
         "Tier 0-3: ~99%",
         "RFC 9449, 9126, 9101, FAPI 2.0, 8705",
         "FAPI 2.0 certification process; JAR+PAR interaction complexity",
         "Write conformance test suite against FAPI 2.0 RP test; target OpenID certification"],
        ["Phase 4\n(Delegation)",
         "Tier 4", "4",
         "Delegation / MCP wiring:\n"
         "• Token exchange grant wired into AS /token\n"
         "• act claim preserved through AS → STS chain\n"
         "• may_act strict enforcement (fix #156)\n"
         "• RFC 9728 PR metadata cross-reference\n"
         "• MCP audience policy (one token per tool server)\n"
         "• RFC 9068 JWT AT act claim profile",
         "Tier 0-4: ~99%",
         "RFC 8693, 9068, 9728, 8707",
         "AS ↔ STS chain latency; may_act circular dependency",
         "Cache STS responses at AS; strict may_act enforcement behind feature flag"],
        ["Phase 5\n(PQC)",
         "Tier 5", "8",
         "Post-Quantum Cryptography:\n"
         "• ML-DSA-65 for ID tokens (extend from AT)\n"
         "• ML-KEM-768 hybrid TLS (X25519MLKEM768)\n"
         "• Hybrid signing (classical + PQC dual)\n"
         "• SIGNING_ALG configurable (fix #160)\n"
         "• dpop_signing_alg_values_supported complete (fix #165)\n"
         "• Oracle risk controls (no decrypt API)\n"
         "• CBOM / SBOM for PQC components",
         "Tier 0-5: ~99% + PQC conformance",
         "FIPS 204, FIPS 203, IANA JOSE, ADR",
         "FIPS 204 IANA JOSE alg registration lag; hybrid negotiation complexity",
         "Use crydi5 / provisional alg IDs internally; document gap; update on IANA finalization"],
    ]

    for i, row_data in enumerate(phases, start=3):
        write_data_row(ws, i, row_data, status_col_idx=None)
        # colour phase cell
        phase_cell = ws.cell(row=i, column=1)
        if "DONE" in str(row_data[0]):
            phase_cell.fill = PatternFill("solid", fgColor=C_IMPL)
        else:
            phase_cell.fill = PatternFill("solid", fgColor=C_MISSING)
        phase_cell.font = Font(bold=True, size=10)

    set_col_widths(ws, [14, 10, 8, 65, 28, 38, 35, 45])


# ── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    output_path = (
        "/Users/Shared/claude/sts-delegate-rs-v2/docs/v2/requirements/"
        "QuAuthz-OAuth21-PQC.xlsx"
    )

    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    build_tab00(wb)
    build_tab01(wb)
    build_tab02(wb)
    build_tab03(wb)
    build_tab04(wb)
    build_tab05(wb)
    build_tab06(wb)
    build_tab07(wb)
    build_tab08(wb)
    build_tab09(wb)
    build_tab10(wb)
    build_tab11(wb)
    build_tab12(wb)

    wb.save(output_path)
    print(f"Saved: {output_path}")

    import os
    size = os.path.getsize(output_path)
    print(f"Size: {size:,} bytes")

    # verify tab names
    names = [ws.title for ws in wb.worksheets]
    print(f"Sheets ({len(names)}): {', '.join(names)}")


if __name__ == "__main__":
    main()
